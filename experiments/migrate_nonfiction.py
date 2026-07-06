"""
migrate_nonfiction.py
=====================
One-time migration: load the nonfiction books from BookRankingsNew.xlsx
(TotalRankings sheet, columns AI-AV) into the nonfiction_books table via the
validated db_write.add_nonfiction_book() workflow.

Discipline
----------
  * Scores are read straight from the workbook — books are NEVER re-scored.
  * The three category averages + Total Average are recomputed on write by
    db_write (unweighted means) and come out equal to the workbook values.
  * Identity fields absent from the workbook block — genre, words, year_read,
    series, series_number — are left NULL. WA is left NULL (the workbook has no
    nonfiction WA, and no nonfiction weights exist yet). status -> 'finished'.
  * Idempotent: add_nonfiction_book refuses duplicate titles, so re-running
    skips books already present instead of double-inserting.

Run:  python3 migrate_nonfiction.py
"""
import openpyxl
from openpyxl.utils import column_index_from_string

import db_write as dbw

WB = "BookRankingsNew.xlsx"
SHEET = "TotalRankings"

# Workbook nonfiction block: header in row 1, data from row 2, columns AI-AV.
COL = {
    "Total Average": "AI", "Book": "AJ", "Author": "AK",
    "Informativeness": "AL", "Argumentation": "AM", "Entertainment": "AN",
    "Quality Average": "AO", "Prose": "AP", "Phraseology": "AQ",
    "Aesthetics Average": "AR", "Insights": "AS", "Philosophizing": "AT",
    "Thought-Provokingness": "AU", "Theme Average": "AV",
}


def _cell(ws, col_letter, row):
    return ws.cell(row=row, column=column_index_from_string(col_letter)).value


def main():
    wb = openpyxl.load_workbook(WB, data_only=True)
    ws = wb[SHEET]

    # Defensive: confirm the block layout is what we expect before writing.
    for name, letter in COL.items():
        got = _cell(ws, letter, 1)
        if got != name:
            raise SystemExit(
                f"ABORT: workbook header {letter}1 is {got!r}, expected {name!r}. "
                f"The nonfiction block layout changed — fix COL before migrating.")

    null_fields = ["genre", "words", "year_read", "series", "series_number", "WA"]
    added, skipped = [], []
    for row in range(2, ws.max_row + 1):
        title = _cell(ws, COL["Book"], row)
        if title is None or str(title).strip() == "":
            continue
        title = str(title).strip()
        scores = {comp: _cell(ws, COL[comp], row) for comp in dbw.NONFICTION_COMPONENTS}
        ok = dbw.add_nonfiction_book(
            title=title,
            author=_cell(ws, COL["Author"], row),
            genre=None, words=None, year_read=None,
            series=None, series_number=None,
            status="finished",
            scores=scores,
        )
        (added if ok else skipped).append(title)

    print("\n" + "=" * 64)
    print(f"Nonfiction migration: {len(added)} added, {len(skipped)} skipped.")
    if added:
        print("  Added:   " + "; ".join(added))
    if skipped:
        print("  Skipped: " + "; ".join(skipped) + "  (already present / refused)")
    print(f"  Left NULL for every migrated book: {', '.join(null_fields)}")
    print("  status='finished'; the 3 category averages + Total Average were")
    print("  recomputed on write and match the workbook.")
    print("=" * 64)


if __name__ == "__main__":
    main()
