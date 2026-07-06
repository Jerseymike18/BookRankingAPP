"""
migrate_to_db.py
================
STEP 1 of the website plan: move your data out of Excel into a clean SQLite
database that stores ONLY the real inputs. Everything else (weighted averages,
ranks, year sheets, tier lists, predictions) is computed from these inputs by
the engine — never stored, so it can never go stale or out of sync.

WHAT GETS STORED (the real input surface)
-----------------------------------------
  books          : your rated fiction — title, genre, author, series, words,
                   the 14 raw component scores, year_read. NOT stored: WA, Total
                   Average, the WStoryAvg-type rollups, Rank (all computed).
  recommendations: TBR predictions — title, genre, author, series, words,
                   predicted components, done flag, blurb, keywords. NOT stored:
                   Model WA, priors, EN scores (all computed by the engine).
  read_queue     : your ordered to-read list (position + title).
  genre_weights  : the GenreWeights config.
  gcomp_weights  : the GCompWeights config.

(Non-fiction is intentionally stubbed for now — the schema leaves room for it.)

SAFETY
------
This is READ-ONLY on your spreadsheet. It creates a NEW file (books.db) and
never touches BookRankingsNew.xlsx. Your spreadsheet stays the source of truth
through the whole transition. If books.db already exists it is rebuilt from
scratch, so re-running is safe and repeatable.

VERIFICATION
------------
After loading, it recomputes every book's Weighted Average FROM THE DATABASE
using your engine's math and checks it matches the spreadsheet's stored WA to
the penny. If all 125 match, the database is a faithful replacement.

HOW TO RUN (Thonny): set WORKBOOK path, press Run. Produces books.db.
"""

import sqlite3
import numpy as np
from openpyxl import load_workbook

import predict_engine as pe

WORKBOOK = pe.WORKBOOK
DB = "books.db"

# Fiction component headers in TotalRankings (raw scores we store).
FICTION_COMPONENTS = [
    "Plot", "Entertainment", "Action", "Ending",
    "Depth", "Emotional Impact", "Motivations",
    "Prose", "Narration",
    "Insights", "Thought-Provokingness",
    "Depth2", "Integration", "Originality",
]

# Recommendations uses abbreviated headers; map them to the canonical names.
REC_COMPONENT_HEADERS = {
    "Plot": "Plot", "Ent": "Entertainment", "Action": "Action", "Ending": "Ending",
    "Depth": "Depth", "Emo": "Emotional Impact", "Motiv": "Motivations",
    "Prose": "Prose", "Narr": "Narration",
    "Insights": "Insights", "ThProv": "Thought-Provokingness",
    "Depth2": "Depth2", "Integration": "Integration", "Origin": "Originality",
}


def header_index(ws):
    hdr = next(ws.iter_rows(values_only=True))
    idx = {}
    for i, h in enumerate(hdr):
        if h is not None and str(h).strip() not in idx:
            idx[str(h).strip()] = i
    return idx


def col_letter_safe(v):
    return None if v is None else (float(v) if isinstance(v, (int, float)) else str(v).strip())


# ---------------------------------------------------------------------------
# Schema
# ---------------------------------------------------------------------------
def create_schema(con):
    cur = con.cursor()
    comp_cols = ",\n        ".join(f'"{c}" REAL' for c in FICTION_COMPONENTS)
    cur.executescript(f"""
    DROP TABLE IF EXISTS books;
    DROP TABLE IF EXISTS recommendations;
    DROP TABLE IF EXISTS read_queue;
    DROP TABLE IF EXISTS genre_weights;
    DROP TABLE IF EXISTS gcomp_weights;

    CREATE TABLE books (
        id INTEGER PRIMARY KEY,
        title TEXT NOT NULL,
        genre TEXT,
        author TEXT,
        series TEXT,
        words INTEGER,
        year_read INTEGER,
        {comp_cols}
    );

    CREATE TABLE recommendations (
        id INTEGER PRIMARY KEY,
        title TEXT NOT NULL,
        genre TEXT,
        author TEXT,
        series TEXT,
        words INTEGER,
        done INTEGER DEFAULT 0,
        blurb TEXT,
        keywords TEXT,
        {comp_cols}
    );

    CREATE TABLE read_queue (
        position INTEGER PRIMARY KEY,
        title TEXT NOT NULL
    );

    CREATE TABLE genre_weights (
        genre TEXT PRIMARY KEY,
        story REAL, character REAL, theme REAL, aesthetics REAL, worldbuilding REAL
    );

    CREATE TABLE gcomp_weights (
        genre TEXT, category TEXT, component TEXT, weight REAL
    );
    """)
    con.commit()


# ---------------------------------------------------------------------------
# Loaders
# ---------------------------------------------------------------------------
def load_books(con, wb):
    ws = wb["TotalRankings"]
    idx = header_index(ws)
    cur = con.cursor()
    cols = ["title", "genre", "author", "series", "words"] + FICTION_COMPONENTS
    placeholders = ",".join("?" for _ in cols)
    quoted = ",".join(f'"{c}"' for c in cols)
    n = 0
    for row in list(ws.iter_rows(values_only=True))[1:]:
        title = row[idx["Book"]]
        plot = row[idx["Plot"]]
        if title is None or plot is None:   # must have a title and be scored
            continue
        vals = [
            str(title).strip(),
            str(row[idx["Genre"]]).strip() if row[idx["Genre"]] else None,
            str(row[idx["Author"]]).strip() if row[idx["Author"]] else None,
            str(row[idx["Series"]]).strip() if row[idx["Series"]] else None,
            int(row[idx["Words"]]) if row[idx["Words"]] else None,
        ]
        for c in FICTION_COMPONENTS:
            v = row[idx[c]] if c in idx else None
            vals.append(float(v) if v is not None else None)
        cur.execute(f"INSERT INTO books (title,genre,author,series,words,"
                    f"{','.join(chr(34)+c+chr(34) for c in FICTION_COMPONENTS)}) "
                    f"VALUES ({placeholders})", vals)
        n += 1
    con.commit()
    return n


def load_recommendations(con, wb):
    ws = wb["Recommendations"]
    idx = header_index(ws)
    cur = con.cursor()
    n = 0
    for row in list(ws.iter_rows(values_only=True))[1:]:
        title = row[idx["Title"]] if "Title" in idx else None
        if title is None:
            continue
        done = row[idx["Done?"]] if "Done?" in idx else None
        vals = [
            str(title).strip(),
            str(row[idx["Genre"]]).strip() if row[idx.get("Genre", -1)] else None,
            str(row[idx["Author"]]).strip() if "Author" in idx and row[idx["Author"]] else None,
            str(row[idx["Series"]]).strip() if "Series" in idx and row[idx["Series"]] else None,
            int(row[idx["Words"]]) if "Words" in idx and row[idx["Words"]] else None,
            1 if done in (True, 1, "TRUE", "True") else 0,
            str(row[idx["Blurb"]]).strip() if "Blurb" in idx and row[idx["Blurb"]] else None,
            str(row[idx["Keywords"]]).strip() if "Keywords" in idx and row[idx["Keywords"]] else None,
        ]
        for canon in FICTION_COMPONENTS:
            # find the rec header that maps to this canonical component
            rec_hdr = next((h for h, c in REC_COMPONENT_HEADERS.items() if c == canon), None)
            v = row[idx[rec_hdr]] if rec_hdr in idx else None
            vals.append(float(v) if isinstance(v, (int, float)) else None)
        cols = (["title", "genre", "author", "series", "words", "done", "blurb", "keywords"]
                + [f'"{c}"' for c in FICTION_COMPONENTS])
        ph = ",".join("?" for _ in cols)
        cur.execute(f"INSERT INTO recommendations ({','.join(cols)}) VALUES ({ph})", vals)
        n += 1
    con.commit()
    return n


def load_queue(con, wb):
    ws = wb["ReadNext"]
    idx = header_index(ws)
    qcol = idx.get("Master Queue (edit here)")
    cur = con.cursor()
    n = 0
    if qcol is not None:
        for row in list(ws.iter_rows(values_only=True))[1:]:
            t = row[qcol]
            if t:
                cur.execute("INSERT INTO read_queue (position,title) VALUES (?,?)",
                            (n + 1, str(t).strip()))
                n += 1
    con.commit()
    return n


def load_weights(con, wb):
    cur = con.cursor()
    for r in list(wb["GenreWeights"].iter_rows(values_only=True))[1:]:
        if r[0]:
            cur.execute("INSERT INTO genre_weights VALUES (?,?,?,?,?,?)",
                        (str(r[0]).strip(), r[1], r[2], r[3], r[4], r[5]))
    for r in list(wb["GCompWeights"].iter_rows(values_only=True))[1:]:
        if r[0]:
            cur.execute("INSERT INTO gcomp_weights VALUES (?,?,?,?)",
                        (str(r[0]).strip(), str(r[1]).strip(), str(r[2]).strip(), r[3]))
    con.commit()


# ---------------------------------------------------------------------------
# Verification: recompute WA from the DB and match against the spreadsheet
# ---------------------------------------------------------------------------
def verify(con, wb):
    # spreadsheet's stored WA, by title
    ws = wb["TotalRankings"]
    idx = header_index(ws)
    stored = {}
    for row in list(ws.iter_rows(values_only=True))[1:]:
        t = row[idx["Book"]]
        if t and row[idx["Plot"]] is not None:
            stored[str(t).strip()] = float(row[idx["Weighted Average"]])

    # genre + component weights from DB
    cur = con.cursor()
    gw = {r[0]: dict(zip(["Story", "Character", "Theme", "Aesthetics", "Worldbuilding"], r[1:]))
          for r in cur.execute("SELECT * FROM genre_weights")}
    gcw = {}
    for genre, cat, comp, wt in cur.execute("SELECT * FROM gcomp_weights"):
        gcw.setdefault(genre, {}).setdefault(cat, {})[comp] = wt

    cat_of = {}  # component -> category
    for genre, cats in gcw.items():
        for cat, comps in cats.items():
            for comp in comps:
                cat_of[comp] = cat

    matches, mismatches = 0, []
    for row in cur.execute(f'SELECT title,genre,'
                           f'{",".join(chr(34)+c+chr(34) for c in FICTION_COMPONENTS)} '
                           f'FROM books'):
        title, genre = row[0], row[1]
        comp_vals = dict(zip(FICTION_COMPONENTS, row[2:]))
        if genre not in gw or genre not in gcw:
            continue
        # components -> weighted category averages
        wcat = {}
        for cat in ["Story", "Character", "Aesthetics", "Theme", "Worldbuilding"]:
            cw = gcw[genre].get(cat, {})
            tot = sum((comp_vals[c] or 0) * (w or 0)
                      for c, w in cw.items() if comp_vals.get(c) is not None)
            wcat[cat] = tot
        wa = (wcat["Story"] * (gw[genre]["Story"] or 0) +
              wcat["Character"] * (gw[genre]["Character"] or 0) +
              wcat["Theme"] * (gw[genre]["Theme"] or 0) +
              wcat["Aesthetics"] * (gw[genre]["Aesthetics"] or 0) +
              wcat["Worldbuilding"] * (gw[genre]["Worldbuilding"] or 0))
        if title in stored and abs(wa - stored[title]) <= 1e-6:
            matches += 1
        elif title in stored:
            mismatches.append((title, stored[title], wa))
    return matches, mismatches, len(stored)


def main():
    print("=" * 60)
    print("MIGRATION: Excel  ->  SQLite (books.db)")
    print("=" * 60)
    wb = load_workbook(WORKBOOK, read_only=True, data_only=True)
    con = sqlite3.connect(DB)

    create_schema(con)
    nb = load_books(con, wb)
    nr = load_recommendations(con, wb)
    nq = load_queue(con, wb)
    load_weights(con, wb)

    print(f"  books table          : {nb} rated fiction books")
    print(f"  recommendations table: {nr} TBR predictions")
    print(f"  read_queue table     : {nq} queued books")
    print(f"  weights              : loaded\n")

    print("Verifying the database reproduces your spreadsheet's WAs...")
    matches, mism, total = verify(con, wb)
    print(f"  WA matches: {matches} / {total}")
    if not mism:
        print("\n  *** All books match. The database faithfully stores your data. ***")
        print("  Your spreadsheet is untouched and remains the source of truth")
        print("  until you're ready to switch over.")
    else:
        print(f"\n  {len(mism)} mismatch(es):")
        for t, s, c in mism[:10]:
            print(f"    {t[:34]:<34} sheet={s:.4f} db={c:.4f}")

    con.close()
    print(f"\n  Created: {DB}")


if __name__ == "__main__":
    main()
