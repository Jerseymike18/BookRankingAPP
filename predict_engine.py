"""
predict_engine.py  (v2 — schema-agnostic)
=========================================
Faithful Python port of your real prediction workflow.

WHAT CHANGED IN v2 (and why)
----------------------------
v1 hardcoded which 19 components existed and which column each lived in
(Plot=8, WStoryAvg=14, ...). When you moved from 19 to 14 components, every
column shifted left and those hardcoded numbers silently pointed at the wrong
data. v2 fixes this permanently:

  * Components are DISCOVERED from the GCompWeights sheet (whatever components
    you currently weight, in whatever categories) — not hardcoded.
  * Columns are found by HEADER NAME, not by position. Add, remove, or reorder
    a component and the engine adapts automatically. You never edit column
    numbers again.

So this same file works for your 14-component layout now, and would work if you
went back to 19 or to any other set tomorrow.

The prediction math (regression -> per-genre bias -> analog blend -> CI -> rank,
plus the Section-7 upstream refinement) is unchanged and still faithful.

HOW TO RUN (Thonny): set WORKBOOK to your real path, edit the example at the
bottom, press Run. Needs pandas, numpy, scipy, openpyxl.
"""

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from scipy import stats

# Point this at your real file. Use the full path that works in Thonny, e.g.
# WORKBOOK = "/Users/michaelpetrides/BookRankings/BookRankingsNew.xlsx"
WORKBOOK = "BookRankingsNew.xlsx"

CATEGORY_WAVG_HEADER = {
    "Story": "WStoryAvg", "Character": "WCharAvg", "Aesthetics": "WAesAvg",
    "Theme": "WThemeAvg", "Worldbuilding": "WWBAvg",
}
REGRESSION_CATS = ["Story", "Character", "Aesthetics", "Theme"]

# ---------------------------------------------------------------------------
# Analog-baseline selection mode (per-component)
# ---------------------------------------------------------------------------
# The per-component analog baseline (see estimate_components) has three nested
# tiers in THIS engine: global -> genre -> author. `author` is the innermost
# (tightest) pool; there is no sub-author "cluster" grouping, so the four-tier
# cluster/author/genre/global scheme from the shrinkage brief collapses to
# three tiers here (author plays the innermost/"cluster" role).
#
#   "hard"   — original behaviour: pick ONE tier by a hard fallback ladder
#              (author mean if >=2 books, else genre mean if >=2, else global).
#              A mean over 2-4 books is trusted as fully as a mean over 10.
#   "shrunk" — empirical-Bayes shrinkage: blend each tier toward its parent,
#              weighted by sample support, so thin tiers get pulled toward the
#              broader (safer) mean and data-rich tiers barely move.
#
# This flag stayed "hard" until the leave-one-out gate in validate_engine.py
# passed, then flipped to "shrunk". Gate result (127-book LOO, K=0.5/0.5):
#   overall WA MAE 0.7740 -> 0.7203 ; small-n bucket 0.6488 -> 0.6326 (better) ;
#   data-rich bucket 0.7565 -> 0.7513 (within +0.03) ; test_engine.py green.
# The shrinkage is strictly UPSTREAM of the genre-bias / DeltaTracker
# corrections — those are unchanged. Re-run `python3 validate_engine.py` to
# reproduce the A/B and the gate.
ANALOG_MODE = "shrunk"    # "hard" | "shrunk"

# Shrinkage strengths ("how many books of support before a tier is trusted on
# its own"). Shared per tier, NOT per component — 126 books cannot support
# per-component tuning without overfitting. Tuned by LOO grid search over the
# brief's grid {0.5,1,2,3,5,8,12}, minimising overall component MAE; the minimum
# is at (0.5, 0.5). LOO MAE keeps dropping marginally below 0.5 (~0.003 WA MAE
# from 0.5 -> 0.1) but that only comes from trusting singleton pools ever harder
# (weight 0.67 at k=0.5 vs 0.91 at k=0.1) — i.e. re-introducing the very
# tiny-pool overfitting this change exists to fix — so 0.5 is the deliberate,
# robust floor. See validate_engine.tune_k(). (research_predict.py uses K_GENRE=6
# / K_AUTHOR=4 for its own, differently-structured WA-level blend.)
K_AUTHOR = 0.5            # author mean -> genre_hat
K_GENRE = 0.5            # genre mean  -> global mean


def _header_map(ws):
    hdr = next(ws.iter_rows(values_only=True))
    out = {}
    for i, h in enumerate(hdr):
        if h is not None:
            name = str(h).strip()
            # Keep the FIRST occurrence only. Several headers (Book, Prose,
            # Insights, ...) repeat in the non-fiction block to the right; the
            # fiction block on the left is the one we want.
            if name not in out:
                out[name] = i
    return out


def discover_schema(path=WORKBOOK):
    """Read GCompWeights to learn the CURRENT components and their categories."""
    wb = load_workbook(path, read_only=True, data_only=True)
    category_components, gcw = {}, {}
    for r in list(wb["GCompWeights"].iter_rows(values_only=True))[1:]:
        if not r[0]:
            continue
        genre, cat, comp, wt = (str(r[0]).strip(), str(r[1]).strip(),
                                str(r[2]).strip(), r[3])
        gcw.setdefault(genre, {}).setdefault(cat, {})[comp] = wt
        category_components.setdefault(cat, [])
        if comp not in category_components[cat]:
            category_components[cat].append(comp)
    return category_components, gcw


def load_everything(path=WORKBOOK):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["TotalRankings"]
    hm = _header_map(ws)
    category_components, gcw = discover_schema(path)
    all_components = [c for comps in category_components.values() for c in comps]

    def col(name):
        if name not in hm:
            raise KeyError(f"Header '{name}' not found in TotalRankings. "
                           f"Available headers start: {list(hm)[:40]}")
        return hm[name]

    c_wa, c_book, c_genre, c_author = (col("Weighted Average"), col("Book"),
                                       col("Genre"), col("Author"))
    wavg_col = {cat: col(h) for cat, h in CATEGORY_WAVG_HEADER.items()}
    comp_col = {comp: col(comp) for comp in all_components}

    recs = []
    for row in list(ws.iter_rows(values_only=True))[1:]:
        book, wa = row[c_book], row[c_wa]
        if book is None or wa is None or row[wavg_col["Story"]] is None:
            continue
        rec = {"Book": str(book).strip(), "WA": float(wa),
               "Genre": str(row[c_genre]).strip() if row[c_genre] else "Unknown",
               "Author": str(row[c_author]).strip() if row[c_author] else "Unknown"}
        for cat, ci in wavg_col.items():
            v = row[ci]
            rec["W" + cat] = float(v) if v is not None else 0.0
        for comp, ci in comp_col.items():
            v = row[ci]
            rec[comp] = float(v) if v is not None else np.nan
        recs.append(rec)
    books = pd.DataFrame(recs)

    gw = {}
    for r in list(wb["GenreWeights"].iter_rows(values_only=True))[1:]:
        if r[0]:
            gw[str(r[0]).strip()] = {"Story": r[1], "Character": r[2], "Theme": r[3],
                                     "Aesthetics": r[4], "Worldbuilding": r[5]}

    books.attrs["category_components"] = category_components
    books.attrs["all_components"] = all_components
    return books, gw, gcw


def components_of(books):
    return books.attrs["all_components"]


def fit_regression(books):
    X = books[["WStory", "WCharacter", "WAesthetics", "WTheme"]].values
    y = books["WA"].values
    X1 = np.column_stack([np.ones(len(X)), X])
    coeffs, *_ = np.linalg.lstsq(X1, y, rcond=None)
    yhat = X1 @ coeffs
    resid = y - yhat
    r2 = 1 - (resid @ resid) / ((y - y.mean()) @ (y - y.mean()))
    resid_sd = float(np.std(resid, ddof=5))
    return coeffs, r2, resid_sd


def regression_wa(coeffs, wstory, wchar, waes, wtheme):
    b0, bs, bc, ba, bt = coeffs
    return b0 + bs * wstory + bc * wchar + ba * waes + bt * wtheme


def genre_bias_and_trust(books, coeffs):
    info = {}
    for g, sub in books.groupby("Genre"):
        preds = regression_wa(coeffs, sub["WStory"], sub["WCharacter"],
                              sub["WAesthetics"], sub["WTheme"])
        bias = float((sub["WA"] - preds).mean())
        n = len(sub)
        info[g] = {"bias": bias, "n": n, "trust": n / (n + 8.0)}
    return info


def fit_upstream(books):
    out = {}
    comps = set(components_of(books))
    drivers = [d for d in ["Plot", "Depth", "Motivations"] if d in comps]
    for target in ["Ending", "Emotional Impact"]:
        if target not in comps or not drivers:
            continue
        d = books[drivers + [target]].dropna()
        X = np.column_stack([np.ones(len(d)), d[drivers].values])
        coef, *_ = np.linalg.lstsq(X, d[target].values, rcond=None)
        out[target] = {"coef": coef, "drivers": drivers}
    return out


def _shrink(n, level_mean, parent_hat, k):
    """One empirical-Bayes shrink step.

    Blend a tier's own mean toward its parent estimate, weighted by how much
    data supports the tier:  (n*level_mean + k*parent_hat) / (n + k).
      * n == 0  -> collapses exactly to parent_hat (the tier contributes
        nothing), so a missing tier needs no special-casing.
      * n >> k  -> converges to level_mean (a data-rich tier barely moves).
    """
    return (n * level_mean + k * parent_hat) / (n + k)


def _shrunk_component_estimates(books, by_author, by_genre, all_components,
                                k_author, k_genre):
    """Per-component nested EB shrinkage: global -> genre_hat -> author_hat.

    author_hat is the analog baseline. Missing tiers have n=0 and collapse to
    their parent (see _shrink), so there is no fallback branching — a book by an
    unseen author shrinks to the genre estimate, and an unseen genre to global.
    n is counted per component (a tier can have full support for Prose but none
    for an optional Worldbuilding component); shrinkage is applied independently
    per component using the shared per-tier constants.
    """
    est = {}
    for comp in all_components:
        gvals = books[comp].dropna()
        global_mean = float(gvals.mean()) if len(gvals) else np.nan
        avals = by_author[comp].dropna()
        grvals = by_genre[comp].dropna()
        n_a, n_g = len(avals), len(grvals)
        author_mean = float(avals.mean()) if n_a else 0.0
        genre_mean = float(grvals.mean()) if n_g else 0.0
        genre_hat = _shrink(n_g, genre_mean, global_mean, k_genre)
        author_hat = _shrink(n_a, author_mean, genre_hat, k_author)
        est[comp] = author_hat
    return est


def estimate_components(books, author, genre, upstream, mode=None,
                        k_author=None, k_genre=None):
    """Estimate the per-component prior for an unread book.

    mode="hard"   -> original fallback ladder (author>=2 else genre>=2 else
                     global), byte-identical to the pre-shrinkage engine.
    mode="shrunk" -> empirical-Bayes shrinkage across the same three tiers.
    mode=None     -> use the module default ANALOG_MODE.

    The Section-7 upstream refinement (Ending / Emotional Impact) is applied
    identically in both modes — it is downstream of the analog baseline, so it
    never confounds the hard-vs-shrunk comparison.
    """
    if mode is None:
        mode = ANALOG_MODE
    all_components = components_of(books)
    by_author = books[books["Author"] == author]
    by_genre = books[books["Genre"] == genre]

    if mode == "shrunk":
        est = _shrunk_component_estimates(
            books, by_author, by_genre, all_components,
            K_AUTHOR if k_author is None else k_author,
            K_GENRE if k_genre is None else k_genre)
        # src label is for the human-facing report only; the shrunk baseline is
        # a blend, so we name the tightest pool that carried any weight.
        if len(by_author):
            src_name, n_src = "author-shrunk", len(by_author)
        elif len(by_genre):
            src_name, n_src = "genre-shrunk", len(by_genre)
        else:
            src_name, n_src = "global-shrunk", len(books)
    else:  # "hard" — original behaviour, unchanged
        if len(by_author) >= 2:
            src_name, src = "author", by_author
        elif len(by_genre) >= 2:
            src_name, src = "genre", by_genre
        else:
            src_name, src = "global", books
        est = {}
        for comp in all_components:
            vals = src[comp].dropna()
            est[comp] = float(vals.mean()) if len(vals) else float(books[comp].dropna().mean())
        n_src = len(src)

    for target, model in upstream.items():
        coef, drivers = model["coef"], model["drivers"]
        if all(d in est for d in drivers):
            pred = coef[0] + sum(coef[k + 1] * est[drivers[k]] for k in range(len(drivers)))
            est[target] = 0.5 * est[target] + 0.5 * float(pred)
    return est, src_name, n_src


def components_to_wcats(comp_values, genre, gcw, books=None):
    wcats = {}
    for cat in REGRESSION_CATS:
        cw = gcw.get(genre, {}).get(cat, {})
        total, used = 0.0, 0.0
        for name, w in cw.items():
            v = comp_values.get(name, np.nan)
            w = w or 0
            if v is None or (isinstance(v, float) and np.isnan(v)):
                continue
            total += float(v) * float(w); used += float(w)
        wcats[cat] = total if used > 0 else 0.0
    return wcats


def predict(title, author, genre, data):
    books, gw, gcw, coeffs, r2, resid_sd, ginfo, upstream = data
    est, src_name, n_src = estimate_components(books, author, genre, upstream)
    wcats = components_to_wcats(est, genre, gcw, books)
    wa_model = regression_wa(coeffs, wcats["Story"], wcats["Character"],
                             wcats["Aesthetics"], wcats["Theme"])
    g = ginfo.get(genre, {"bias": 0.0, "n": 0, "trust": 0.0})
    wa_corrected = wa_model + g["bias"]
    analog_wa = books[books["Author"] == author]["WA"].dropna()
    if len(analog_wa) < 2:
        analog_wa = books[books["Genre"] == genre]["WA"].dropna()
    analog_mean = float(analog_wa.mean()) if len(analog_wa) else wa_corrected
    trust = g["trust"]
    wa_final = trust * wa_corrected + (1 - trust) * analog_mean
    half = 1.645 * resid_sd
    ci_low, ci_high = wa_final - half, wa_final + half
    rank = int((books["WA"] > wa_final).sum() + 1)
    rank_hi = int((books["WA"] > ci_low).sum() + 1)
    rank_lo = int((books["WA"] > ci_high).sum() + 1)
    return {"title": title, "author": author, "genre": genre, "est": est,
            "wcats": wcats, "wa_model": wa_model, "bias": g["bias"],
            "trust": trust, "analog_mean": analog_mean, "wa_final": wa_final,
            "ci": (ci_low, ci_high), "rank": rank, "rank_range": (rank_lo, rank_hi),
            "src": src_name, "n_src": n_src, "n_genre": g["n"],
            "total": len(books), "r2": r2, "resid_sd": resid_sd}


def report(p):
    print("=" * 64)
    print(f"PREDICTION  —  {p['title']}")
    print(f"            {p['author']}  |  {p['genre']}")
    print("=" * 64)
    print(f"Component estimate source: {p['src']} (n={p['n_src']})   [un-researched prior]\n")
    print("Estimated weighted category averages (regression inputs):")
    for c in REGRESSION_CATS:
        print(f"   W{c:<10} {p['wcats'][c]:.2f}")
    print()
    print(f"  Regression point estimate : {p['wa_model']:.2f}   (model R²={p['r2']:.3f})")
    print(f"  Genre bias correction     : {p['bias']:+.3f}")
    print(f"  Model/analog blend (trust : {p['trust']:.2f}, analog mean {p['analog_mean']:.2f})")
    print(f"  ─────────────────────────────────────")
    print(f"  PREDICTED WA              : {p['wa_final']:.2f}")
    print(f"  90% CI                    : [{p['ci'][0]:.2f}, {p['ci'][1]:.2f}]")
    print(f"  Predicted Total-Avg Rank  : ~{p['rank']} of {p['total']}  "
          f"(range {p['rank_range'][0]}–{p['rank_range'][1]})")
    if p["n_genre"] < 5:
        print(f"  ** Thin genre (n={p['n_genre']}): leaning on analogs; treat as rough. **")
    print()
    print("NOTE: component scores are an autonomous estimate, not researched.")


def build(source="db", path=WORKBOOK):
    if source == "db":
        import db_loader  # lazy: db_loader imports this module
        books, gw, gcw = db_loader.load_from_db()
    else:
        books, gw, gcw = load_everything(path)
    coeffs, r2, resid_sd = fit_regression(books)
    ginfo = genre_bias_and_trust(books, coeffs)
    upstream = fit_upstream(books)
    return books, gw, gcw, coeffs, r2, resid_sd, ginfo, upstream


if __name__ == "__main__":
    data = build()
    print(f"Loaded {len(data[0])} books.")
    print(f"Current components ({len(components_of(data[0]))}): "
          f"{', '.join(components_of(data[0]))}\n")
    report(predict("The Republic of Thieves", "Scott Lynch", "Epic Fantasy", data))
