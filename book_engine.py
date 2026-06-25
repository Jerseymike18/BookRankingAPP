"""
book_engine.py
==============
A Python port of Michael's book-rating prediction engine, with an honest
validation harness.

WHAT THIS DOES, IN PLAIN LANGUAGE
---------------------------------
Your spreadsheet predicts how much you'll like a book by looking at similar
books you've already read (same author, then same "cluster", then same genre).
This script does the same thing in Python, but then it does something the
spreadsheet can't easily do: it HIDES each book one at a time, predicts it as
if it were unseen, and checks how close the prediction was. That's called
"leave-one-out" validation, and it's the honest way to measure whether the
engine actually predicts -- or just memorizes books it has already seen.

It compares three prediction strategies (the same fallback ladder you use):
    1. AUTHOR mean  -- average score of other books by the same author
    2. CLUSTER mean -- average within an author-or-genre cluster
    3. GENRE mean   -- average score within the same genre

...and reports, for each, how far off it was on average (MAE = Mean Absolute
Error, in score points on your 0-10 scale). Lower = better.

HOW TO RUN
----------
    python book_engine.py

Requires: pandas, numpy, openpyxl  (install with: pip install pandas numpy openpyxl)
"""

import numpy as np
import pandas as pd
from openpyxl import load_workbook

WORKBOOK = "BookRankingsNew.xlsx"

# ---------------------------------------------------------------------------
# STEP 1: Load the data out of Excel into a clean table
# ---------------------------------------------------------------------------
# We read TotalRankings (your source of truth). Each row is a book with an
# overall Weighted Average (WA) score plus its genre, author, and series.

# Column positions in TotalRankings (0-indexed), confirmed from the sheet:
COL = {
    "WA": 2,          # C: Weighted Average  (the headline score we predict)
    "TotalAvg": 1,    # B: Total Average
    "Book": 3,        # D
    "Genre": 4,       # E
    "Author": 5,      # F
    "Series": 6,      # G
    "Words": 7,       # H
}
# The 19 component sub-scores live in these columns:
COMPONENT_COLS = {
    "Plot": 8, "Entertainment": 9, "Pacing": 10, "Action": 11, "Ending": 12,
    "Depth": 15, "EmotionalImpact": 16, "Motivations": 17, "MultiDim": 18,
    "Prose": 21, "Narration": 22, "Phraseology": 23,
    "Insights": 26, "Philosophizing": 27, "ThoughtProvoking": 28,
    "WBDepth": 31, "WBIntegration": 32, "WBOriginality": 33, "WBConsistency": 34,
}


def load_books(path=WORKBOOK):
    """Read TotalRankings into a pandas DataFrame of scored fiction books."""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["TotalRankings"]
    records = []
    for row in list(ws.iter_rows(values_only=True))[1:]:  # skip header
        book = row[COL["Book"]]
        wa = row[COL["WA"]]
        plot = row[COMPONENT_COLS["Plot"]]
        # A valid scored book has a title, a WA, and at least a Plot score.
        if book is None or wa is None or plot is None:
            continue
        rec = {
            "Book": str(book).strip(),
            "WA": float(wa),
            "Genre": str(row[COL["Genre"]]).strip() if row[COL["Genre"]] else "Unknown",
            "Author": str(row[COL["Author"]]).strip() if row[COL["Author"]] else "Unknown",
            "Series": str(row[COL["Series"]]).strip() if row[COL["Series"]] else "Standalone",
        }
        for name, idx in COMPONENT_COLS.items():
            v = row[idx]
            rec[name] = float(v) if v is not None else np.nan
        records.append(rec)
    df = pd.DataFrame(records)
    return df


# ---------------------------------------------------------------------------
# STEP 2: The three predictors (your fallback ladder), each as leave-one-out
# ---------------------------------------------------------------------------
# For each predictor we predict book i using ONLY the other books -- never
# book i itself. This is what makes the accuracy number honest.

def loo_predict(df, group_col, min_n=2):
    """
    Leave-one-out prediction using the mean WA of other books in the same group.

    group_col : "Author", "Genre", or a custom cluster column.
    min_n     : need at least this many OTHER books in the group to predict;
                otherwise we return NaN (can't predict -> not counted).

    Returns a Series of predictions aligned to df's index.
    """
    preds = pd.Series(np.nan, index=df.index)
    for grp, idxs in df.groupby(group_col).groups.items():
        idxs = list(idxs)
        if len(idxs) < min_n + 1:
            continue  # not enough peers to leave one out
        wa = df.loc[idxs, "WA"]
        total = wa.sum()
        for i in idxs:
            others_mean = (total - wa[i]) / (len(idxs) - 1)
            preds[i] = others_mean
    return preds


def build_cluster(df):
    """
    Recreate your 'cluster' idea: an author gets their own cluster if they have
    enough books (>=3), otherwise they fall into a genre-level bucket. This
    mirrors your CrossValidation Section 3 (EF-Erikson, EF-Genre (other EF), ...).
    """
    author_counts = df["Author"].value_counts()
    big_authors = set(author_counts[author_counts >= 3].index)
    def cluster_of(r):
        if r["Author"] in big_authors:
            return f"{r['Genre'][:2].upper()}-{r['Author']}"
        return f"{r['Genre']}-genre"
    return df.apply(cluster_of, axis=1)


# ---------------------------------------------------------------------------
# STEP 3: Scoring helpers
# ---------------------------------------------------------------------------
def mae(pred, actual):
    mask = pred.notna()
    if mask.sum() == 0:
        return np.nan, 0
    return float(np.abs(pred[mask] - actual[mask]).mean()), int(mask.sum())


def within(pred, actual, tol):
    mask = pred.notna()
    if mask.sum() == 0:
        return np.nan
    return float((np.abs(pred[mask] - actual[mask]) <= tol).mean())


# ---------------------------------------------------------------------------
# STEP 4: Run it all and print an honest report
# ---------------------------------------------------------------------------
def main():
    df = load_books()
    n = len(df)
    print("=" * 68)
    print("BOOK PREDICTION ENGINE -- HONEST VALIDATION REPORT")
    print("=" * 68)
    print(f"Loaded {n} scored books across {df['Genre'].nunique()} genres.")
    print(f"Overall WA: mean {df['WA'].mean():.2f}, "
          f"min {df['WA'].min():.2f}, max {df['WA'].max():.2f}, "
          f"std {df['WA'].std():.2f}")
    print()
    print("Baseline to beat: 'just guess the overall average for every book'.")
    naive_mae = float(np.abs(df["WA"] - df["WA"].mean()).mean())
    print(f"  Naive (global mean) MAE = {naive_mae:.3f}  "
          f"<- any real predictor must beat this")
    print()

    df["Cluster"] = build_cluster(df)

    strategies = {
        "Author mean": loo_predict(df, "Author"),
        "Cluster mean": loo_predict(df, "Cluster"),
        "Genre mean": loo_predict(df, "Genre"),
    }

    print("-" * 68)
    print(f"{'Strategy':<16}{'MAE':>8}{'n predicted':>14}"
          f"{'within 0.5':>12}{'within 1.0':>12}")
    print("-" * 68)
    for name, preds in strategies.items():
        m, cnt = mae(preds, df["WA"])
        w05 = within(preds, df["WA"], 0.5)
        w10 = within(preds, df["WA"], 1.0)
        print(f"{name:<16}{m:>8.3f}{cnt:>14}{w05:>11.0%}{w10:>11.0%}")
    print("-" * 68)
    print()

    # Per-genre breakdown: where is the engine strong vs. weak?
    print("PER-GENRE accuracy (Cluster strategy) -- and the small-sample warning")
    print("-" * 68)
    print(f"{'Genre':<30}{'n':>4}{'MAE':>9}{'verdict':>22}")
    print("-" * 68)
    cluster_preds = strategies["Cluster mean"]
    for g, idxs in df.groupby("Genre").groups.items():
        idxs = list(idxs)
        sub_pred = cluster_preds[idxs]
        m, cnt = mae(sub_pred, df.loc[idxs, "WA"])
        if cnt == 0:
            verdict = "can't predict (too few)"
            mstr = "  --  "
        else:
            mstr = f"{m:.3f}"
            if cnt < 5:
                verdict = "TRUST CAUTIOUSLY (n<5)"
            elif m < 0.6:
                verdict = "strong"
            elif m < 0.9:
                verdict = "okay"
            else:
                verdict = "weak"
        print(f"{g:<30}{len(idxs):>4}{mstr:>9}{verdict:>22}")
    print("-" * 68)
    print()
    print("HOW TO READ THIS")
    print("  * MAE = average miss, in score points. 0.6 means typically off")
    print("    by about half a point on your 0-10 scale.")
    print("  * Any strategy must beat the naive baseline above to be 'real'.")
    print("  * Genres with n<5 are flagged: those accuracy numbers are built on")
    print("    so few books that they're mostly noise -- do NOT trust them as")
    print("    evidence the engine works. The honest signal lives in your big")
    print("    genres (Epic Fantasy, Sci-Fi, Science Fantasy).")


if __name__ == "__main__":
    main()
