"""
walkforward.py
==============
Chronological walk-forward backtest of the researched-prediction engine.

THE QUESTION THIS ANSWERS
-------------------------
"What would the engine have predicted for each book on the day I started it,
using only what was known then?" -- an HONEST accuracy baseline that future
engine features must beat, plus the raw dataset for a future track-record page.

This is stricter than the leave-one-out sweep in validate_engine.py. LOO trains
on the other 126 books INCLUDING future ones; walk-forward trains only on books
read BEFORE the held-out book, in real reading order (the Timeline sheet).

WHAT IT IS (and is NOT)
-----------------------
  * It CALLS the read-only engine; it never modifies prediction math.
    predict_engine / db_loader / reresearch_and_measure / research_predict are
    used exactly as the live Predict page uses them (research vector -> optional
    correlation-smoothing -> author+genre correction -> WA roll-up).
  * It NEVER writes books.db. Every fold is an in-memory filter of the books
    DataFrame -- no scratch DB file needed, because the researched-prediction
    functions are all parameterised by the `books` frame + `cache` dict (there
    is no hardcoded DB read inside the prediction path). See PHASE-0 NOTES.
  * ZERO API SPEND, structurally. The harness reads the richer-prompt cache
    (llm_scores_richer.json) as a plain dict and calls only the pure correction
    functions. It monkeypatches anthropic.Anthropic to raise, so any accidental
    client construction fails loudly rather than spending a cent. A book with no
    usable cache entry is logged SKIPPED_NO_CACHE -- never researched.
  * DETERMINISTIC. Same DB + caches -> byte-identical folds artifact (sorted
    keys, fixed float rounding, no timestamps in the folds file).

THREE VARIANTS PER FOLD (all run, all from cache)
-------------------------------------------------
  raw    -- research vector rolled straight to WA: no correlation smoothing, no
            author+genre correction. "How good is the uncalibrated grounded LLM
            research?" Pool-independent by construction.
  honest -- full pipeline (smooth + author_genre correction), every trainable
            piece (correction pairs, smoothing models, resid_sd, rank) fit on
            the PAST-ONLY pool (positions 1..t-1). THE walk-forward baseline.
  leaky  -- same pipeline, but fit on the FULL library (today's config). Labeled
            leaky because the correction saw future books. "How good is today's
            engine config", not "what was knowable then".

PHASE-0 NOTES (leakage + architecture facts this harness relies on)
-------------------------------------------------------------------
  * The DeltaTracker `component_corrections` table is RETIRED (active row has
    all-zero constants, blend 0) AND is never read by the prediction path, so it
    does not enter any variant. The correction that actually shapes predictions
    is reresearch_and_measure.correct_book (method "author_genre").
  * Research-cache vectors embed post-publication reception -- an ACCEPTED
    hindsight caveat (a walk-forward run cannot un-know a book's reputation).
  * The interval recorded per variant is the engine's own +/-1.645*resid_sd
    (a ~90% normal interval, exactly what correct_and_predict emits), NOT the
    served density-bucketed conformal interval (that needs a full-LOO residual
    table and is out of scope here).

HOW TO RUN
----------
    python3 walkforward.py                 # run all folds + write report
    python3 walkforward.py --report-only   # rebuild report from existing folds
    python3 walkforward.py --check-determinism   # prove two runs are identical
    python3 walkforward.py --burn-in 15    # min pool size before evaluating

Artifacts land in validation/ (NOT a static-snapshot input -- see README).
"""

import argparse
import hashlib
import json
import os
import sqlite3
import db_backend
import subprocess

import numpy as np
import pandas as pd

# Read-only engine + the exact live-Predict glue. Importing research_predict
# pulls in `anthropic`, but NO client is constructed and NO network call is made
# here -- we call only its pure functions and guard the client below.
import predict_engine as pe
import db_loader
import reresearch_and_measure as rm
import research_predict as rp

ROOT = os.path.dirname(os.path.abspath(__file__))
LIVE = rm.LIVE                       # canonical 14 components, reference order
WB = set(rm.WB)                      # the 3 worldbuilding comps (0.0 sentinel)
OUT_DIR = os.path.join(ROOT, "validation")
FOLDS_FILE = "walkforward_folds.jsonl"
META_FILE = "walkforward_meta.json"
REPORT_FILE = "walkforward_report.md"
ROLLING_FILE = "walkforward_rolling_mae.json"

BURN_IN_DEFAULT = 15                 # min training-pool size before we evaluate
ROLL_WINDOW = 15                     # rolling-MAE window (folds)
NOMINAL_COVERAGE = 0.90              # +/-1.645*resid_sd is a 90% normal interval
VARIANTS = ("raw", "honest", "leaky")

# Verbatim into the results metadata (brief Phase 0.2). "neutralised" == a
# past-only pool filter removes the future-information leak for that input.
LEAKAGE_INVENTORY = {
    "author_genre_correction_pairs": "neutralised in honest (pool), LEAKY in leaky (full library)",
    "correlation_smoothing_models": "neutralised in honest (pool), LEAKY in leaky (full library)",
    "resid_sd_for_interval": "neutralised in honest (pool), LEAKY in leaky (full library)",
    "rank_over_library": "neutralised in honest (pool); leaky ranks over full library",
    "genre_and_component_weights": "config, not learned from the book set -- no leakage",
    "component_corrections_deltatracker": "retired to zero AND unwired -- enters no variant",
    "research_cache_vector": "ACCEPTED hindsight caveat -- embeds post-publication reception",
}


# ---------------------------------------------------------------------------
# Zero-API structural guard
# ---------------------------------------------------------------------------
def _install_no_api_guard():
    """Make the harness structurally incapable of spending tokens: replace the
    Anthropic client constructor so ANY accidental client build raises. The
    prediction functions we call (correct_and_predict / build_corr_models /
    correct_book / _wa_from_components) never construct a client, so this is a
    belt-and-braces backstop, not a behavioural change."""
    import anthropic

    def _blocked(*_a, **_k):
        raise RuntimeError(
            "walkforward.py is zero-spend: Anthropic client construction is "
            "blocked. A cache miss must be logged SKIPPED_NO_CACHE, not researched.")

    anthropic.Anthropic = _blocked


# ---------------------------------------------------------------------------
# Determinism helpers
# ---------------------------------------------------------------------------
def _r(x):
    """Round to a fixed precision and coerce numpy -> python float, so two runs
    serialise byte-identically. None passes through."""
    if x is None:
        return None
    if isinstance(x, (int, np.integer)) and not isinstance(x, bool):
        return int(x)
    return round(float(x), 6)


def _rd(d):
    """Round every value of a component dict deterministically."""
    return {c: _r(d.get(c)) for c in LIVE}


def _git_head():
    try:
        return subprocess.check_output(
            ["git", "rev-parse", "HEAD"], cwd=ROOT,
            stderr=subprocess.DEVNULL).decode().strip()
    except Exception:
        return None


def _engine_hash():
    """Content hash of every file whose code determines a prediction. Changes if
    the engine or its glue changes, so a stale folds artifact is detectable."""
    h = hashlib.sha256()
    for name in ("predict_engine.py", "db_loader.py",
                 "reresearch_and_measure.py", "research_predict.py"):
        try:
            with open(os.path.join(ROOT, name), "rb") as fh:
                h.update(fh.read())
        except OSError:
            h.update(b"\0MISSING\0")
    return "sha256:" + h.hexdigest()[:16]


def _active_correction_version(db_path):
    """Read (read-only) the active component_corrections version, for the meta
    header. Returns a short descriptor even though this layer enters no variant
    (it is retired + unwired) -- recording it documents that fact per run."""
    try:
        uri = "file:" + os.path.abspath(db_path) + "?mode=ro"
        con = db_backend.connect(uri, uri=True)
        row = con.execute(
            "SELECT version, decision FROM component_corrections "
            "WHERE active=1 LIMIT 1").fetchone()
        con.close()
        if row:
            return {"version": row[0], "decision": row[1],
                    "applied_in_engine": False}
    except Exception:
        pass
    return {"version": None, "decision": None, "applied_in_engine": False}


# ---------------------------------------------------------------------------
# Ordering (Timeline read order + place-last for un-timelined books)
# ---------------------------------------------------------------------------
def _timeline_order(xlsx_path):
    """Extract (read_number, title) pairs from the Timeline sheet's per-book
    table (the sub-table with #, Book, Author, ... columns). We trust Timeline
    ONLY for order; all book metadata comes from the DB (source of truth)."""
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["Timeline"]
    rows = list(ws.iter_rows(values_only=True))

    # Header row = the one that carries an 'Author' label (only the per-book
    # table has it). Locate #/Book relative to Author within that row.
    hdr_idx = next((i for i, r in enumerate(rows)
                    if any(str(c).strip() == "Author" for c in r if c is not None)), None)
    if hdr_idx is None:
        raise RuntimeError("Timeline: could not find the per-book header row.")
    hdr = [str(c).strip() if c is not None else "" for c in rows[hdr_idx]]
    auth_col = hdr.index("Author")
    book_col = max(i for i, v in enumerate(hdr) if v == "Book" and i < auth_col)
    num_col = max(i for i, v in enumerate(hdr) if v == "#" and i < book_col)

    out = []
    for r in rows[hdr_idx + 1:]:
        title = r[book_col] if book_col < len(r) else None
        if title in (None, ""):
            continue
        num = r[num_col] if num_col < len(r) else None
        out.append((num, str(title).strip()))
    return out


def build_order(books, xlsx_path):
    """Return (ordered_positions, skips).

    ordered_positions: list of dicts, one per rated fiction book, in read order,
    each carrying position (1-based) + DB metadata. Fiction books present in the
    Timeline come first in Timeline order (renumbered 1..K after dropping the
    interleaved nonfiction rows); the owner-approved rule for rated books ABSENT
    from the Timeline (recent additions with no recorded order) is to place them
    LAST, ordered by (year_read, title) for determinism.

    skips: books that cannot be placed (should be empty under current data)."""
    db_by_title = {row["Book"]: row for _, row in books.iterrows()}
    db_titles = set(db_by_title)

    timeline = _timeline_order(xlsx_path)
    # Timeline order, keeping only rated FICTION titles (nonfiction rows in the
    # sheet simply don't match the fiction `books` table and are dropped).
    seen = set()
    ordered_titles = []
    for _num, title in sorted(timeline, key=lambda t: (t[0] is None, t[0])):
        if title in db_titles and title not in seen:
            seen.add(title)
            ordered_titles.append(title)

    # Place-last: rated fiction books absent from the Timeline, by (year_read, title).
    missing = sorted(
        (t for t in db_titles if t not in seen),
        key=lambda t: (db_by_title[t]["Year"] if db_by_title[t]["Year"] is not None
                       else 9999, t))

    positions = []
    for pos, title in enumerate(ordered_titles + missing, start=1):
        row = db_by_title[title]
        positions.append({
            "position": pos,
            "title": title,
            "author": row["Author"],
            "genre": row["Genre"],
            "series": row["Series"] or None,
            "series_number": _series_number(title, books),
            "year_read": int(row["Year"]) if row["Year"] is not None else None,
            "in_timeline": title in seen,
        })
    return positions, []


def _series_number(title, books):
    """series_number lives in the DB but db_loader does not surface it; read it
    straight from books.db (read-only) once, cached on the function."""
    cache = getattr(_series_number, "_cache", None)
    if cache is None:
        cache = {}
        try:
            uri = "file:" + os.path.abspath(db_loader.DB) + "?mode=ro"
            con = db_backend.connect(uri, uri=True)
            for t, n in con.execute("SELECT title, series_number FROM books"):
                cache[str(t).strip()] = n
            con.close()
        except Exception:
            pass
        _series_number._cache = cache
    return cache.get(title)


# ---------------------------------------------------------------------------
# Variant prediction
# ---------------------------------------------------------------------------
def _errors(pred_components, pred_wa, actual_components, actual_wa):
    """Signed + absolute error for WA and every component."""
    comp_signed, comp_abs = {}, {}
    for c in LIVE:
        a = actual_components.get(c)
        p = pred_components.get(c)
        if a is None or p is None or (isinstance(a, float) and np.isnan(a)):
            comp_signed[c], comp_abs[c] = None, None
        else:
            comp_signed[c] = _r(p - a)
            comp_abs[c] = _r(abs(p - a))
    return {
        "wa_signed_error": _r(pred_wa - actual_wa),
        "wa_abs_error": _r(abs(pred_wa - actual_wa)),
        "component_signed_error": comp_signed,
        "component_abs_error": comp_abs,
    }


def _variant_raw(raw_scores, genre, gw, gcw, resid_sd, actual_components, actual_wa):
    """RAW: research vector -> WA, no smoothing, no correction. Pool-independent
    components; borrows the honest pool resid_sd only to state an interval."""
    wa = rp._wa_from_components(raw_scores, genre, gw, gcw)
    half = 1.645 * resid_sd
    rec = {
        "wa": _r(wa), "components": _rd(raw_scores),
        "ci_low": _r(wa - half), "ci_high": _r(wa + half),
        "ci_inside": bool(wa - half <= actual_wa <= wa + half),
        "resid_sd": _r(resid_sd), "rank": None, "rank_total": None,
        "n_author": None, "n_genre": None, "analog_src": "none",
    }
    rec.update(_errors(raw_scores, wa, actual_components, actual_wa))
    return rec


def _variant_corrected(title, author, genre, raw_scores, conf, books_train,
                       resid_sd, corr_models, gw, gcw, cache,
                       actual_components, actual_wa):
    """HONEST or LEAKY (identical code; the caller decides the training frame).
    Runs the exact live pipeline: correlation-smooth -> author+genre correct ->
    WA roll-up, via research_predict.correct_and_predict."""
    res = rp.correct_and_predict(
        title, author, genre, dict(raw_scores), conf, resid_sd,
        books_train, gw, gcw, cache, corr_models=corr_models)
    wa = res["wa"]
    ci_low, ci_high = res["ci"]
    n_author, n_genre = res["n_author"], res["n_genre"]
    rec = {
        "wa": _r(wa), "components": _rd(res["scores"]),
        "ci_low": _r(ci_low), "ci_high": _r(ci_high),
        "ci_inside": bool(ci_low <= actual_wa <= ci_high),
        "resid_sd": _r(resid_sd), "rank": res["rank"], "rank_total": res["total"],
        "n_author": n_author, "n_genre": n_genre,
        "analog_src": ("author" if n_author > 0
                       else "genre" if n_genre > 0 else "global"),
    }
    rec.update(_errors(res["scores"], wa, actual_components, actual_wa))
    return rec


def run_folds(books, gw, gcw, cache, order, burn_in):
    """Walk the reading order, predicting each book at position t (> burn_in)
    from the past-only pool. Returns (fold_records, skip_records)."""
    # LEAKY config is the SAME for every fold ("today's engine"): full-library
    # smoothing models + resid_sd, computed once. correct_and_predict excludes
    # the target row from the correction training internally.
    resid_sd_full = pe.fit_regression(books)[2]
    corr_models_full = rp.build_corr_models(books, cache)

    folds, skips = [], []
    title_to_row = {row["Book"]: row for _, row in books.iterrows()}

    for entry in order:
        pos = entry["position"]
        title = entry["title"]

        if title not in cache or not isinstance(cache[title].get("scores"), dict):
            skips.append({"skip": True, "position": pos, "title": title,
                          "reason": "SKIPPED_NO_CACHE"})
            continue
        if pos <= burn_in:
            skips.append({"skip": True, "position": pos, "title": title,
                          "reason": "BURN_IN"})
            continue

        raw_scores = {c: float(cache[title]["scores"][c])
                      for c in LIVE if c in cache[title]["scores"]}
        if len(raw_scores) != len(LIVE):
            skips.append({"skip": True, "position": pos, "title": title,
                          "reason": "SKIPPED_NO_CACHE"})
            continue
        conf = cache[title].get("conf", "?")

        row = title_to_row[title]
        actual_wa = float(row["WA"])
        actual_components = {c: (float(row[c]) if row[c] is not None
                                 and not (isinstance(row[c], float) and np.isnan(row[c]))
                                 else None) for c in LIVE}
        author, genre = entry["author"], entry["genre"]

        # Past-only pool = books at strictly-earlier reading positions.
        pool_titles = [e["title"] for e in order if e["position"] < pos]
        books_pool = books[books["Book"].isin(pool_titles)]

        resid_sd_pool = pe.fit_regression(books_pool)[2]
        corr_models_pool = rp.build_corr_models(books_pool, cache)

        variants = {
            "raw": _variant_raw(raw_scores, genre, gw, gcw, resid_sd_pool,
                                actual_components, actual_wa),
            "honest": _variant_corrected(
                title, author, genre, raw_scores, conf, books_pool,
                resid_sd_pool, corr_models_pool, gw, gcw, cache,
                actual_components, actual_wa),
            "leaky": _variant_corrected(
                title, author, genre, raw_scores, conf, books,
                resid_sd_full, corr_models_full, gw, gcw, cache,
                actual_components, actual_wa),
        }

        folds.append({
            "position": pos, "title": title, "author": author, "genre": genre,
            "series": entry["series"], "series_number": entry["series_number"],
            "year_read": entry["year_read"], "in_timeline": entry["in_timeline"],
            "pool_size": int(len(books_pool)), "cache_key": title,
            "actual_wa": _r(actual_wa), "actual_components": _rd(actual_components),
            "variants": variants,
        })

    folds.sort(key=lambda f: f["position"])
    skips.sort(key=lambda s: (s["position"] if s["position"] is not None else 0,
                              s["title"]))
    return folds, skips


# ---------------------------------------------------------------------------
# Artifact I/O
# ---------------------------------------------------------------------------
def _serialise_folds(folds, skips):
    """Deterministic JSONL text: fold records (by position) then skip records.
    No timestamps -> byte-identical across runs on the same DB + caches."""
    lines = [json.dumps(f, sort_keys=True) for f in folds]
    lines += [json.dumps(s, sort_keys=True) for s in skips]
    return "\n".join(lines) + "\n"


def write_artifacts(folds, skips, books, cache, order, burn_in, out_dir):
    os.makedirs(out_dir, exist_ok=True)
    with open(os.path.join(out_dir, FOLDS_FILE), "w") as fh:
        fh.write(_serialise_folds(folds, skips))

    import datetime
    skip_counts = {}
    for s in skips:
        skip_counts[s["reason"]] = skip_counts.get(s["reason"], 0) + 1
    meta = {
        "generated_at": datetime.datetime.now().astimezone().isoformat(timespec="seconds"),
        "git_head": _git_head(),
        "engine_hash": _engine_hash(),
        "correction_version_active": _active_correction_version(db_loader.DB),
        "burn_in": burn_in,
        "n_books_total": len(order),
        "n_folds_evaluated": len(folds),
        "n_skipped": len(skips),
        "skip_reasons": skip_counts,
        "n_in_cache": sum(1 for e in order if e["title"] in cache),
        "components": LIVE,
        "variants": {
            "raw": "research vector -> WA (no smoothing, no correction); pool-independent",
            "honest": "smooth + author_genre correction fit on PAST-ONLY pool (the walk-forward baseline)",
            "leaky": "smooth + author_genre correction fit on FULL library (today's config; leaky)",
        },
        "nominal_interval_coverage": NOMINAL_COVERAGE,
        "interval_note": "per-variant interval is the engine's +/-1.645*resid_sd, not the served conformal interval",
        "leakage_inventory": LEAKAGE_INVENTORY,
        "caveats": [
            "research-cache vectors embed post-publication reception (hindsight) -- accepted",
            "leaky variant's correction saw future books -- labeled leaky, not a knowable-then number",
            "3 rated books absent from Timeline are placed last by (year_read, title) per owner decision",
        ],
    }
    with open(os.path.join(out_dir, META_FILE), "w") as fh:
        json.dump(meta, fh, indent=2, sort_keys=True)
        fh.write("\n")
    return meta


def load_folds(out_dir):
    """Read back the folds artifact -> (folds, skips), so --report-only works
    standalone from a prior run."""
    folds, skips = [], []
    path = os.path.join(out_dir, FOLDS_FILE)
    with open(path) as fh:
        for line in fh:
            line = line.strip()
            if not line:
                continue
            rec = json.loads(line)
            (skips if rec.get("skip") else folds).append(rec)
    folds.sort(key=lambda f: f["position"])
    return folds, skips


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
def _load_inputs():
    books, gw, gcw = db_loader.load_from_db()
    cache = rp.load_cache()
    return books, gw, gcw, cache


def do_run(burn_in, out_dir, xlsx_path):
    _install_no_api_guard()
    books, gw, gcw, cache = _load_inputs()
    order, _ = build_order(books, xlsx_path)
    folds, skips = run_folds(books, gw, gcw, cache, order, burn_in)
    meta = write_artifacts(folds, skips, books, cache, order, burn_in, out_dir)
    print(f"Walk-forward: {meta['n_folds_evaluated']} folds evaluated, "
          f"{meta['n_skipped']} skipped ({meta['skip_reasons']}).")
    print(f"  wrote {os.path.join(out_dir, FOLDS_FILE)}")
    print(f"  wrote {os.path.join(out_dir, META_FILE)}")
    return folds, skips


def do_check_determinism(burn_in, xlsx_path):
    _install_no_api_guard()
    books, gw, gcw, cache = _load_inputs()
    order, _ = build_order(books, xlsx_path)
    a = _serialise_folds(*run_folds(books, gw, gcw, cache, order, burn_in))
    b = _serialise_folds(*run_folds(books, gw, gcw, cache, order, burn_in))
    ha, hb = hashlib.sha256(a.encode()).hexdigest(), hashlib.sha256(b.encode()).hexdigest()
    print(f"run A sha256: {ha}")
    print(f"run B sha256: {hb}")
    if a == b:
        print("DETERMINISM: PASS (two runs byte-identical)")
        return True
    print("DETERMINISM: FAIL")
    return False


def main():
    ap = argparse.ArgumentParser(
        description="Chronological walk-forward backtest of the researched "
                    "prediction engine (zero-spend, read-only).")
    ap.add_argument("--burn-in", type=int, default=BURN_IN_DEFAULT,
                    help=f"min training-pool size before evaluating (default {BURN_IN_DEFAULT}).")
    ap.add_argument("--out-dir", default=OUT_DIR, help="artifact directory.")
    ap.add_argument("--xlsx", default=os.path.join(ROOT, "BookRankingsNew.xlsx"),
                    help="workbook holding the Timeline read order.")
    ap.add_argument("--report-only", action="store_true",
                    help="rebuild the report from an existing folds artifact.")
    ap.add_argument("--check-determinism", action="store_true",
                    help="run the folds twice and assert byte-identical output.")
    args = ap.parse_args()

    if args.check_determinism:
        raise SystemExit(0 if do_check_determinism(args.burn_in, args.xlsx) else 1)

    if args.report_only:
        _maybe_build_report(args.out_dir, required=True)
        return

    do_run(args.burn_in, args.out_dir, args.xlsx)
    _maybe_build_report(args.out_dir, required=False)


def _maybe_build_report(out_dir, required):
    """Build the markdown report if the report module is present. It ships in a
    later commit than the core harness, so the harness stays runnable without it."""
    try:
        import walkforward_report as wr
    except ImportError:
        msg = ("walkforward_report.py not found -- run the harness first, then "
               "the report module (added in the report commit).")
        if required:
            raise SystemExit(msg)
        print(f"  (skipped report: {msg})")
        return
    wr.build_report(out_dir)


if __name__ == "__main__":
    main()
