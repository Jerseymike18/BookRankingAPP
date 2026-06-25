"""
backfill_year_read.py
=====================
One-time backfill: populate the empty `year_read` column in books.db from the
2025BooksRead / 2026BooksRead sheets of the spreadsheet, so year-views and the
Timeline dashboard can be built.

SAFE: reads the spreadsheet read-only, writes only the year_read column in the
DB, makes a timestamped backup of books.db first, and verifies every book got
tagged before finishing. Re-runnable.

Matching is case-insensitive (the spreadsheet has "Speaker for the Dead", the DB
has "Speaker For The Dead" — same book).

HOW TO RUN (Thonny): press Run.
"""

import sqlite3
import shutil
import datetime as dt
from openpyxl import load_workbook
import predict_engine as pe

DB = "books.db"
WORKBOOK = pe.WORKBOOK


def year_books(wb, sheet):
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    return set(str(r[3]).strip() for r in rows[1:] if r[3] and r[8] is not None)


def main():
    wb = load_workbook(WORKBOOK, read_only=True, data_only=True)
    b2026 = year_books(wb, "2026BooksRead")
    b2025 = year_books(wb, "2025BooksRead")
    # case-insensitive lookup: lower-title -> year
    year_of = {}
    for t in b2025:
        year_of[t.lower()] = 2025
    for t in b2026:
        year_of[t.lower()] = 2026   # 2026 wins if somehow in both

    # backup
    stamp = dt.datetime.now().strftime("%Y%m%d-%H%M%S")
    shutil.copy2(DB, f"{DB}.backup-{stamp}")
    print(f"  (backup saved: {DB}.backup-{stamp})")

    con = sqlite3.connect(DB)
    rows = con.execute("SELECT id, title FROM books").fetchall()
    tagged, untagged = 0, []
    for bid, title in rows:
        y = year_of.get(title.strip().lower())
        if y:
            con.execute("UPDATE books SET year_read=? WHERE id=?", (y, bid))
            tagged += 1
        else:
            untagged.append(title)
    con.commit()

    print(f"  Tagged {tagged}/{len(rows)} books with year_read.")
    if untagged:
        print(f"  NOT tagged ({len(untagged)}) — not found in either year sheet:")
        for t in untagged:
            print(f"    {t}")
        print("  (These may be books you added via the app after the spreadsheet")
        print("   was archived — set their year via the app going forward.)")
    else:
        print("  *** Every book got a year. Backfill complete. ***")

    # verify
    counts = con.execute("SELECT year_read, COUNT(*) FROM books GROUP BY year_read").fetchall()
    print(f"  Year distribution now: {counts}")
    con.close()


if __name__ == "__main__":
    main()
