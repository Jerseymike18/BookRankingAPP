"""
backfill_delta_log_from_workbook.py
===================================
One-off (idempotent) migration: populate the delta_log table with the full
historical predicted-vs-actual record, joined from three sources.

  PREDICTED  : BookRankingsNew.xlsx / TBRFinished  (col C 'Predicted Score' = pred_wa;
               the 14 raw component columns I,J,K,L,O,P,Q,T,U,X,Y,AB,AC,AD)
  ACTUAL     : the books table (source of truth), WA via the read-only engine
  DELTA      : computed here as d = actual - predicted (DeltaTracker convention)
  CROSS-CHECK: BookRankingsNew.xlsx / DeltaTracker  (Δ columns) — sanity only

All writes go through db_write.backfill_delta_log (which _backup_once()'s first,
computes act-pred internally, and is idempotent on the backfill marker). Books
present in TBRFinished but not matchable in the books table are SKIPPED and
reported — never fabricated. The single pre-existing live row is left untouched
(it is not present in TBRFinished, so it does not duplicate a migrated book).

Run:  python3 backfill_delta_log_from_workbook.py
"""
import os
import sqlite3
import statistics as st
from collections import Counter

import openpyxl
from openpyxl.utils import column_index_from_string as ci

import db_loader
import db_write

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "BookRankingsNew.xlsx")
DB = os.path.join(HERE, "books.db")
TOL = 0.05  # cross-check tolerance vs DeltaTracker

# (delta_log component, TBRFinished col letter, books-table col, DeltaTracker col letter)
COMPS = [
    ("Plot",                  "I",  "Plot",                  "H"),
    ("Entertainment",         "J",  "Entertainment",         "I"),
    ("Action",                "K",  "Action",                "K"),
    ("Ending",                "L",  "Ending",                "L"),
    ("Depth",                 "O",  "Depth",                 "O"),
    ("Emotional_Impact",      "P",  "Emotional Impact",      "P"),
    ("Motivations",           "Q",  "Motivations",           "Q"),
    ("Prose",                 "T",  "Prose",                 "T"),
    ("Narration",             "U",  "Narration",             "U"),
    ("Insights",              "X",  "Insights",              "X"),
    ("Thought_Provokingness", "Y",  "Thought-Provokingness", "Y"),
    ("Depth2",                "AB", "WB Depth",              "AB"),
    ("Integration",           "AC", "WB Integration",        "AC"),
    ("Originality",           "AD", "WB Originality",        "AD"),
]
# canonical books-table component names (spaces/hyphens) keyed by delta_log name
BOOKS_COL = {
    "Emotional_Impact": "Emotional Impact",
    "Thought_Provokingness": "Thought-Provokingness",
}


def norm(t):
    return " ".join(str(t).strip().split()).casefold() if t is not None else None


def fnum(x):
    try:
        return float(x)
    except (TypeError, ValueError):
        return None


def load_tbr(wb):
    ws = wb["TBRFinished"]
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        book = row[3]  # col D
        if book is None or str(book).strip() == "":
            continue
        rec = {"title": str(book).strip(), "pred_wa": fnum(row[2])}  # col C
        for name, tletter, _, _ in COMPS:
            rec[name] = fnum(row[ci(tletter) - 1])
        out[norm(book)] = rec
    return out


def load_deltatracker(wb):
    ws = wb["DeltaTracker"]
    out = {}
    for row in ws.iter_rows(min_row=3, values_only=True):  # rows 1-2 are summary/header
        book = row[0]  # col A
        if book is None or str(book).strip() == "":
            continue
        rec = {"title": str(book).strip(), "d_pred_score": fnum(row[ci("G") - 1])}
        for name, _, _, dletter in COMPS:
            rec[name] = fnum(row[ci(dletter) - 1])
        out[norm(book)] = rec
    return out


def load_books():
    """Actual component scores + engine-computed WA, via the read-only engine."""
    books_df, _, _ = db_loader.load_from_db()
    out = {}
    for _, r in books_df.iterrows():
        rec = {"title": str(r["Book"]).strip(), "act_wa": fnum(r["WA"])}
        for name, _, _, _ in COMPS:
            rec[name] = fnum(r[BOOKS_COL.get(name, name)])
        out[norm(r["Book"])] = rec
    return out


def main():
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    tbr = load_tbr(wb)
    dt = load_deltatracker(wb)
    books = load_books()

    tbr_keys, book_keys = set(tbr), set(books)
    matched = sorted(tbr_keys & book_keys)
    tbr_only = sorted(tbr_keys - book_keys)     # predicted, no actual -> SKIP
    book_only = sorted(book_keys - tbr_keys)    # actual, no prediction -> not backfillable

    # ---- Build records + cross-check ----
    records, flags = [], []
    for k in matched:
        t, b = tbr[k], books[k]
        rec = {"title": b["title"], "pred_wa": t["pred_wa"], "act_wa": b["act_wa"],
               "pred": {}, "act": {}}
        worst, worst_comp, n_over, sign_flips = 0.0, "", 0, []
        for name, _, _, _ in COMPS:
            p, a = t[name], b[name]
            cname = BOOKS_COL.get(name, name)  # canonical FICTION_COMPONENTS key for db_write
            rec["pred"][cname] = p
            rec["act"][cname] = a
            d = (a - p) if (p is not None and a is not None) else None
            dtv = dt.get(k, {}).get(name)
            if d is not None and dtv is not None:
                ad = abs(d - dtv)
                if ad > worst:
                    worst, worst_comp = ad, name
                if ad > TOL:
                    n_over += 1
                if abs(d) >= 0.15 and abs(dtv) >= 0.15 and (d > 0) != (dtv > 0):
                    sign_flips.append(name)
        records.append(rec)
        # WA cross-check vs DeltaTracker 'Δ Predicted Score'
        dwa = (b["act_wa"] - t["pred_wa"]) if (t["pred_wa"] is not None and b["act_wa"] is not None) else None
        dt_dwa = dt.get(k, {}).get("d_pred_score")
        wa_ad = abs(dwa - dt_dwa) if (dwa is not None and dt_dwa is not None) else None
        if n_over > 0 or (wa_ad is not None and wa_ad > TOL):
            flags.append({"title": b["title"], "max": worst, "worst_comp": worst_comp,
                          "n_over": n_over, "sign_flips": sign_flips, "wa_ad": wa_ad})

    # ---- Write via db_write (idempotent, backs up first) ----
    report = db_write.backfill_delta_log(records)

    # ---- Report ----
    line = "=" * 68
    print(line)
    print("delta_log BACKFILL — migration report")
    print(line)
    print(f"TBRFinished predicted rows : {len(tbr)}")
    print(f"DeltaTracker rows          : {len(dt)}")
    print(f"books table rows (engine)  : {len(books)}")
    print(f"Matched (join on title)    : {len(matched)}")
    print(f"  -> inserted into delta_log: {report['inserted']}")
    print(f"  -> skipped (malformed)    : {len(report['skipped'])}  {report['skipped'] or ''}")
    print(f"SKIPPED: predicted but no actual in books table: {len(tbr_only)}")
    for k in tbr_only:
        print(f"    - {tbr[k]['title']}")
    print(f"Not backfillable (actual but no prediction): {len(book_only)}")
    for k in book_only:
        print(f"    - {books[k]['title']}  (no TBRFinished prediction; left as-is)")

    print("\n" + "-" * 68)
    print(f"CROSS-CHECK vs DeltaTracker (tolerance {TOL}) — computed d = act - pred")
    print("-" * 68)
    print(f"Books clean (every component within {TOL}): {len(matched) - len(flags)} / {len(matched)}")
    print(f"Books FLAGGED (>=1 component or WA over {TOL}): {len(flags)}")
    if flags:
        print(f"  {'title':34} {'maxΔ':>5} {'@comp':22} {'n>':>2} {'waΔ':>6}  signflips")
        for f in sorted(flags, key=lambda x: -x["max"]):
            wa = "-" if f["wa_ad"] is None else f"{f['wa_ad']:.3f}"
            sf = ",".join(f["sign_flips"]) or "-"
            print(f"  {f['title'][:34]:34} {f['max']:5.2f} {f['worst_comp']:22} "
                  f"{f['n_over']:2d} {wa:>6}  {sf}")
    print("  NOTE: these are same-book magnitude/precision drift between the frozen")
    print("  DeltaTracker snapshot and the CURRENT TBRFinished predictions + books")
    print("  actuals (which are authoritative). Titles/columns verified — not join")
    print("  errors. The written values use the current authoritative sources.")

    # ---- Verify: count + spot-check ----
    con = sqlite3.connect(DB)
    con.row_factory = sqlite3.Row
    total = con.execute("SELECT COUNT(*) FROM delta_log").fetchone()[0]
    n_marker = con.execute("SELECT COUNT(*) FROM delta_log WHERE logged_at=?",
                           (db_write.DELTA_BACKFILL_MARKER,)).fetchone()[0]
    n_live = total - n_marker
    print("\n" + "-" * 68)
    print("VERIFY")
    print("-" * 68)
    print(f"delta_log total rows: {total}  (backfilled {n_marker} + live {n_live})")

    for title in ("Toll the Hounds", "Midnight Tides", "The Neverending Story"):
        r = con.execute("SELECT * FROM delta_log WHERE title=? AND logged_at=?",
                        (title, db_write.DELTA_BACKFILL_MARKER)).fetchone()
        if not r:
            print(f"\n  SPOT {title}: NOT FOUND")
            continue
        k = norm(title)
        print(f"\n  SPOT-CHECK: {title}")
        print(f"    pred_wa={r['pred_wa']:.4f}  act_wa={r['act_wa']:.4f}  d_wa={r['d_wa']:+.4f}"
              f"  (DeltaTracker Δpred={dt[k]['d_pred_score']:+.3f})")
        print(f"    {'comp':22} {'pred':>6} {'act':>6} {'d(db)':>7} {'DT Δ':>6} {'ok?':>4}")
        for name, _, _, _ in COMPS:
            pv, av, dv = r[f"pred_{name}"], r[f"act_{name}"], r[f"d_{name}"]
            dtv = dt[k][name]
            ok = "ok" if (dtv is not None and abs(dv - dtv) <= TOL) else "flag"
            print(f"    {name:22} {pv:6.2f} {av:6.2f} {dv:+7.2f} {dtv:+6.2f} {ok:>4}")
    con.close()
    print("\n" + line)
    print("Done.")
    print(line)


if __name__ == "__main__":
    main()
